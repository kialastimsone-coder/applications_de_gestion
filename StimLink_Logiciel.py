import os
import pathlib
import random
import smtplib
import tempfile

import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter import messagebox


# Accessibilité
def accesibilite():
    global fen
    fen = tk.Tk()
    fen.config(background='gray20')
    fen.geometry('545x300')
    fen.minsize(545, 300)
    fen.maxsize(545, 300)
    fen.title('StimLink Logiciel')
    fen.iconbitmap('Design StimLink 2-1.ico')
    labelprincipale = Label(fen, text="BIENVENUE A STIMLOG DATABASE", font=('calibri', 25, 'bold'), bg='gray20'
                            , fg='gold', bd=40)
    labelprincipale.pack(fill=X)

    detail_frame = LabelFrame(fen, text='Se connecter', font=('calibri', 18, 'bold'),
                              fg='yellow', bg='gray20', relief=GROOVE, bd=6)

    detail_frame.pack(fill=X)

    global email
    email = tk.StringVar()
    Champ = tk.Entry(detail_frame, textvariable=email, highlightthickness=2, bg='gray20', fg='white',
                     font=('calibri', 14, 'bold'), width=30)
    Champlabel1 = tk.Label(detail_frame, text='Adresse Email', font=('calibri', 14, 'bold'),
                           bg='gray20', fg='white')
    Champlabel1.grid(row=0, column=0, sticky='w')
    Champ.grid(row=0, column=1, padx=15, pady=5)

    global motdepasse
    motdepasse = tk.StringVar()
    Champ2 = tk.Entry(detail_frame, textvariable=motdepasse, highlightthickness=2, show='*',
                      bg='gray20', fg='white', font=('calibri', 14, 'bold'), width=30)
    Champlabel2 = tk.Label(detail_frame, text='Mot de passe', font=('calibri', 14, 'bold'),
                           bg='gray20', fg='white')
    Champlabel2.grid(row=1, column=0, sticky='w')
    Champ2.grid(row=1, column=1, padx=20, pady=5)

    Valider = Button(detail_frame, text='Valider', font=('Arial', 11, 'bold'), bd=3, width=8, command=connexion)
    Valider.grid(row=2, column=1, padx=20, pady=5)

    fen.mainloop()


matnumber = random.randint(100, 999)


def connexion():
    if email.get() == 'Stimsonekiala@gmail.com' and motdepasse.get() == 'sivi1706':
        messagebox.showinfo('Réussi', 'Succefuly ID')
        grande_fonction()

    elif email.get() == '' or motdepasse.get() == '':
        messagebox.showerror('Echec', 'Inserez votre Email et votre Mot de passe !')

    elif email.get() != 'Stimsonekiala@gmail.com' and motdepasse.get() != 'sivi1706':
        messagebox.showerror('Echec', 'Adresse email saisi et le mot de passe sont incorrectes !\nVeillez réessayez...')
        email.set('')
        motdepasse.set('')

    elif email.get() == 'Stimsonekiala@gmail.com' and motdepasse.get() != 'sivi1706':
        messagebox.showerror('Echec', 'Mot de passe incorrecte !\nVeillez réessayez...')
        motdepasse.set('')

    elif email.get() != 'Stimsonekiala@gmail.com' and motdepasse.get() == 'sivi1706':
        messagebox.showerror('Echec', 'Adresse email incorrecte !\nVeillez réessayez...')
        email.set('')

    else:
        messagebox.showerror('Erroné', 'ID incorrecte')
        email.set('')
        motdepasse.set('')


# Grande fonction
def grande_fonction():
    destroy()

    def Base_de_donnees():
        os.system("start:Database.xlsx")
        file = pathlib.Path("Database.xlsx")
        if file.exists():
            pass
        else:
            file = Workbook()
            sheet = file.active
            sheet['A1'] = "Matricule"
            sheet['B1'] = "Nom et post-nom"
            sheet['C1'] = "Classe"
            sheet['D1'] = "Cours"
            sheet['E1'] = "Interro 1"
            sheet['F1'] = "Interro 2"
            sheet['G1'] = "Interro 3"
            sheet['H1'] = "Interro 4"
            sheet['I1'] = "Interro 5"
            sheet['J1'] = "TP 1"
            sheet['K1'] = "TP 2"
            sheet['L1'] = "TP 3"
            sheet['M1'] = "TP 4"
            sheet['N1'] = "TP 5"
            sheet['O1'] = "TJ 1"
            sheet['P1'] = "TJ 2"
            sheet['Q1'] = "TJ 3"
            sheet['R1'] = "TJ 4"
            sheet['S1'] = "TJ 5"
            sheet['T1'] = "IG 1"
            sheet['U1'] = "IG 2"
            sheet['V1'] = "IG 3"
            sheet['W1'] = "Examen 1"
            sheet['X1'] = "Examen 2"

            file.save("Database.xlsx")

    # LES FONCTIONS
    def recherche():
        for i in os.listdir('LISTE/'):
            if i.split('.')[0] == entrymatricule.get():
                f = open(f'LISTE/{i}', 'r')
                textresult.delete(1.0, END)
                for data in f:
                    textresult.insert(END, data)
                f.close()
                break
        else:
            messagebox.showerror('Erreur', 'N° matricule incorrect')
            entrymatricule.delete(0, END)

    def imprimer():
        if textresult.get(1.0, END) == '\n':
            messagebox.showerror('Erreur', "Il n'y a aucun résultat à imprimer !")
        else:
            file = tempfile.mktemp('.txt')
            open(file, 'w').write(textresult.get(1.0, END))
            os.startfile(file, 'print')

    if not os.path.exists('LISTE'):
        os.mkdir('LISTE')

    def email():
        def envoi():
            try:
                OB = smtplib.SMTP('stmp.gmail.com', 587)
                OB.starttls()
                OB.login(xenderEntry.get(), motdepasseEntry.get())
                messages = messagetext.get(1.0, END)
                OB.sendmail(xenderEntry.get(), entrydestinateur.get(), messages)
                OB.quit()
                messagebox.showinfo('Success', 'Mail envoyé avec succès...', parent=root1)
                root1.destroy()

            except:
                messagebox.showerror('Erreur', "Vérifiez la connexion internet\nEt/ou vérifier si votre émail et "
                                               "mot de passe sont correcte, puis réessayez", parent=root1)
                root1.destroy()

        if textresult.get(1.0, END) == '\n':
            messagebox.showerror('Erreur', "Pas de document à envoyer...")
        else:
            root1 = Toplevel()
            root1.grab_set()
            root1.title("Envoi des résultats par Email")
            root1.config(bg="gray20")
            root1.resizable(0, 0)

            xenderframe = LabelFrame(root1, text="TRANSFERT", font=('calibri', 17, 'bold'), bd=5, bg='gray20',
                                     fg='white')
            xenderframe.grid(row=0, column=0)

            gmailframe = Label(xenderframe, text="Email", font=('calibri', 13, 'bold'), bg='gray20',
                               fg='white')
            gmailframe.grid(row=0, column=0, sticky='w')

            xenderEntry = Entry(xenderframe, font=('calibri', 13, 'bold'), bd=2, width=23, relief=RIDGE)
            xenderEntry.grid(row=0, column=1, padx=10, pady=8)

            gmailframe = Label(xenderframe, text="Mot de passe", font=('calibri', 13, 'bold'), bg='gray20',
                               fg='white')
            gmailframe.grid(row=1, column=0, sticky='w')

            motdepasseEntry = Entry(xenderframe, font=('calibri', 13, 'bold'), bd=2, width=23, relief=RIDGE, show='*')
            motdepasseEntry.grid(row=1, column=1, padx=10, pady=8, sticky='w')

            destinateurframe = LabelFrame(root1, text="DESTINATEUR", font=('calibri', 17, 'bold'), bd=5, bg='gray20',
                                          fg='white')
            destinateurframe.grid(row=1, column=0, padx=40, pady=20)

            destframe = Label(destinateurframe, text="Adresse Email", font=('calibri', 13, 'bold'), bg='gray20',
                              fg='white')
            destframe.grid(row=0, column=0, sticky='w', padx=10, pady=8)

            entrydestinateur = Entry(destinateurframe, font=('calibri', 13, 'bold'), bd=2, width=23, relief=RIDGE)
            entrydestinateur.grid(row=0, column=1, padx=10, pady=8)

            messagelabel = Label(destinateurframe, text="Message", font=('calibri', 13, 'bold'), bg='gray20',
                                 fg='white')
            messagelabel.grid(row=1, column=0, padx=10, pady=8)

            messagetext = Text(destinateurframe, font=('calibri', 13, 'bold'), bd=2, relief=SUNKEN, width=42,
                               height=11)
            messagetext.grid(row=2, column=0, columnspan=2)
            messagetext.delete(1.0, END)
            messagetext.insert(END, textresult.get(1.0, END).replace('=', '').replace('-', '').replace('~', ''))

            envoyer = Button(root1, text='Envoyer', font=('calibri', 13, 'bold'), width=13, command=envoi)
            envoyer.grid(row=2, column=0, pady=20)

            root1.mainloop()

    def somme():

        global matnumber, sommeInterro
        global sommeTp
        global sommeTj
        global sommeEXAIG
        global POURCENT

        sommeInterro = 0
        I1 = int(interro1Entry.get())
        I2 = int(interro2Entry.get())
        I3 = int(interro3Entry.get())
        I4 = int(interro4Entry.get())
        I5 = int(interro5Entry.get())
        if I1 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif I2 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif I3 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif I4 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif I5 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        else:
            sommeInterro = round((I1 + I2 + I3 + I4 + I5) / 5, 2)

        totalInterrogation.delete(0, END)
        totalInterrogation.insert(0, str(sommeInterro))

        sommeTp = 0
        Tp1 = int(TP1Entry.get())
        Tp2 = int(TP2Entry.get())
        Tp3 = int(TP3Entry.get())
        Tp4 = int(TP4Entry.get())
        Tp5 = int(TP5Entry.get())
        if Tp1 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif Tp2 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif Tp3 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif Tp4 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif Tp5 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        else:
            sommeTp = round((Tp1 + Tp2 + Tp3 + Tp4 + Tp5) / 5, 2)

        totaltp.delete(0, END)
        totaltp.insert(0, str(sommeTp))

        sommeTj = 0
        Tj1 = int(TJ1Entry.get())
        Tj2 = int(TJ2Entry.get())
        Tj3 = int(TJ3Entry.get())
        Tj4 = int(TJ4Entry.get())
        Tj5 = int(TJ5Entry.get())
        if Tj1 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif Tj2 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif Tj3 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif Tj4 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif Tj5 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        else:
            sommeTj = round((Tj1 + Tj2 + Tj3 + Tj4 + Tj5) / 5, 2)

        totaltj.delete(0, END)
        totaltj.insert(0, str(sommeTj))

        sommeEXAIG = 0
        EXAIG1 = int(IG1Entry.get())
        EXAIG2 = int(IG2Entry.get())
        EXAIG3 = int(IG3Entry.get())
        EXAIG4 = int(Exa1Entry.get())
        EXAIG5 = int(Exa2Entry.get())

        if EXAIG1 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif EXAIG2 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif EXAIG3 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif EXAIG4 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        elif EXAIG5 > 10:
            messagebox.showerror('Erreur', 'La valeur ne peut pas depasée 10\nMerci de réessayez !')
        else:
            sommeEXAIG = round(EXAIG1 + EXAIG2 + EXAIG3 + EXAIG4 + EXAIG5 / 5, 2)

        moyenne = round((sommeInterro + sommeTp + sommeTj + sommeEXAIG) / 4, 2)
        frame.delete(0, END)
        frame.insert(0, str(moyenne))

        POURCENT = 0
        if moyenne == 0.0:
            pourcentage.insert(0, END)

        else:
            POURCENT = round(float(I1 + I2 + I3 + I4 + I5 + Tp1 + Tp2 + Tp3 + Tp4 + Tp5 +
                                   Tj1 + Tj2 + Tj3 + Tj4 + Tj5 + sommeEXAIG) * 50 / 100, 2)

        pourcentage.delete(0, END)
        pourcentage.insert(0, f'{str(POURCENT)} %')

        # GRADE
        Grades = Grade
        if moyenne == 0.0:
            Grades.delete(0, END)
            Grades.insert(0, '')

        elif moyenne < 8.0:
            Grades.delete(0, END)
            Grades.insert(0, 'L')

        elif 8.0 <= moyenne < 10.0:
            Grades.delete(0, END)
            Grades.insert(0, 'F')

        elif 10.0 <= moyenne < 12.0:
            Grades.delete(0, END)
            Grades.insert(0, 'G')

        elif 12.0 <= moyenne < 14.0:
            Grades.delete(0, END)
            Grades.insert(0, 'B')

        elif 14.0 <= moyenne < 16.0:
            Grades.delete(0, END)
            Grades.insert(0, 'E')

        elif 16.0 <= moyenne < 18.0:
            Grades.delete(0, END)
            Grades.insert(0, 'D')

        elif moyenne >= 18.0:
            Grades.delete(0, END)
            Grades.insert(0, 'E')

    def sauvegarder():
        global matnumber
        matricule = matnumber
        noms = entryname.get()
        classe = entrynumber.get()
        cours = entryCOURS.get()
        Interro1 = interro1Entry.get()
        Interro2 = interro2Entry.get()
        Interro3 = interro3Entry.get()
        Interro4 = interro4Entry.get()
        Interro5 = interro5Entry.get()
        tp1 = TP1Entry.get()
        tp2 = TP2Entry.get()
        tp3 = TP3Entry.get()
        tp4 = TP4Entry.get()
        tp5 = TP5Entry.get()
        tj1 = TJ1Entry.get()
        tj2 = TJ2Entry.get()
        tj3 = TJ3Entry.get()
        tj4 = TJ4Entry.get()
        tj5 = TJ5Entry.get()
        ig1 = IG1Entry.get()
        ig2 = IG2Entry.get()
        ig3 = IG3Entry.get()
        examen1 = Exa1Entry.get()
        examen2 = Exa2Entry.get()

        if entryname.get() == '' or entrynumber.get() == '' or entryCOURS.get() == '' or Grade == '':
            messagebox.showerror('Erreur', "Le document ne pas prêt pour l'enregistrement !")

        else:
            result = messagebox.askyesno('Sauvegarder', 'Voulez-vous enregistrer ?')
            if result:
                newenreg = textresult.get(1.0, END)
                file = open(f'LISTE/{matnumber}.txt', 'w')
                file.write(newenreg)
                file.close()
                messagebox.showinfo('Opération réussie', f' {entryname.get()} est enregistrer avec succès !')
                matnumber = random.randint(100, 999)

                file = openpyxl.load_workbook('Database.xlsx')
                sheet = file.active
                sheet.cell(column=1, row=sheet.max_row + 1, value=int(matricule))
                sheet.cell(column=2, row=sheet.max_row, value=noms)
                sheet.cell(column=3, row=sheet.max_row, value=classe)
                sheet.cell(column=4, row=sheet.max_row, value=cours)
                sheet.cell(column=5, row=sheet.max_row, value=int(Interro1))
                sheet.cell(column=6, row=sheet.max_row, value=int(Interro2))
                sheet.cell(column=7, row=sheet.max_row, value=int(Interro3))
                sheet.cell(column=8, row=sheet.max_row, value=int(Interro4))
                sheet.cell(column=9, row=sheet.max_row, value=int(Interro5))
                sheet.cell(column=10, row=sheet.max_row, value=int(tp1))
                sheet.cell(column=11, row=sheet.max_row, value=int(tp2))
                sheet.cell(column=12, row=sheet.max_row, value=int(tp3))
                sheet.cell(column=13, row=sheet.max_row, value=int(tp4))
                sheet.cell(column=14, row=sheet.max_row, value=int(tp5))
                sheet.cell(column=15, row=sheet.max_row, value=int(tj1))
                sheet.cell(column=16, row=sheet.max_row, value=int(tj2))
                sheet.cell(column=17, row=sheet.max_row, value=int(tj3))
                sheet.cell(column=18, row=sheet.max_row, value=int(tj4))
                sheet.cell(column=19, row=sheet.max_row, value=int(tj5))
                sheet.cell(column=20, row=sheet.max_row, value=int(ig1))
                sheet.cell(column=21, row=sheet.max_row, value=int(ig2))
                sheet.cell(column=22, row=sheet.max_row, value=int(ig3))
                sheet.cell(column=23, row=sheet.max_row, value=int(examen1))
                sheet.cell(column=24, row=sheet.max_row, value=int(examen2))

                file.save("Database.xlsx")

        entryname.delete(0, END)
        entrynumber.delete(0, END)
        entryCOURS.delete(0, END)
        entrymatricule.delete(0, END)

        interro1Entry.delete(0, END)
        interro2Entry.delete(0, END)
        interro3Entry.delete(0, END)
        interro4Entry.delete(0, END)
        interro5Entry.delete(0, END)
        TP1Entry.delete(0, END)
        TP2Entry.delete(0, END)
        TP3Entry.delete(0, END)
        TP4Entry.delete(0, END)
        TP5Entry.delete(0, END)
        TJ1Entry.delete(0, END)
        TJ2Entry.delete(0, END)
        TJ3Entry.delete(0, END)
        TJ4Entry.delete(0, END)
        TJ5Entry.delete(0, END)
        IG1Entry.delete(0, END)
        IG2Entry.delete(0, END)
        IG3Entry.delete(0, END)
        Exa1Entry.delete(0, END)
        Exa2Entry.delete(0, END)

        interro1Entry.insert(0, 0)
        interro2Entry.insert(0, 0)
        interro3Entry.insert(0, 0)
        interro4Entry.insert(0, 0)
        interro5Entry.insert(0, 0)
        TP1Entry.insert(0, 0)
        TP2Entry.insert(0, 0)
        TP3Entry.insert(0, 0)
        TP4Entry.insert(0, 0)
        TP5Entry.insert(0, 0)
        TJ1Entry.insert(0, 0)
        TJ2Entry.insert(0, 0)
        TJ3Entry.insert(0, 0)
        TJ4Entry.insert(0, 0)
        TJ5Entry.insert(0, 0)
        IG1Entry.insert(0, 0)
        IG2Entry.insert(0, 0)
        IG3Entry.insert(0, 0)
        Exa1Entry.insert(0, 0)
        Exa2Entry.insert(0, 0)

        totalInterrogation.delete(0, END)
        totaltp.delete(0, END)
        totaltj.delete(0, END)
        frame.delete(0, END)
        pourcentage.delete(0, END)
        Grade.delete(0, END)

        textresult.delete(1.0, END)

    def apercu():
        if entryname.get() == '' or entrynumber.get() == '' or entryCOURS.get() == '':
            messagebox.showerror('Erreur', 'Veillez remplir les champs svp !')

        elif totalInterrogation.get() == '' and totaltp.get() == '' and totaltj.get() == '':
            messagebox.showerror('Vide', 'Les totaux sont vides !')

        elif totalInterrogation.get() == '0.0' and totaltp.get() == '0.0' and totaltj.get() == '0.0':
            messagebox.showerror('No insert', "Le résultat n'est pas prêt !")

        else:
            textresult.insert(END, '          ******** STIMLOG RESULT ********\n')
            textresult.insert(END, f'\n                                     Matricule Number : {matnumber}')
            textresult.insert(END, f'\nNames    : {entryname.get()}')
            textresult.insert(END, f'\nClass    : {entrynumber.get()}')
            textresult.insert(END, f'\nCourse   : {entryCOURS.get()}\n')
            textresult.insert(END, f'\n==========================================================')
            textresult.insert(END, f'\nInterrogation 1 : {interro1Entry.get()}/10')
            textresult.insert(END, f'\nInterrogation 2 : {interro2Entry.get()}/10')
            textresult.insert(END, f'\nInterrogation 3 : {interro3Entry.get()}/10')
            textresult.insert(END, f'\nInterrogation 4 : {interro4Entry.get()}/10')
            textresult.insert(END, f'\nInterrogation 5 : {interro5Entry.get()}/10')
            textresult.insert(END, f'\n                                         - - - - - - - - -')
            textresult.insert(END, f'\n                                         Moyenne : {totalInterrogation.get()}/10')
            textresult.insert(END, f'\n                                         - - - - - - - - -')

            textresult.insert(END, f'\n\n==========================================================')
            textresult.insert(END, f'\nTravail pratique 1 : {TP1Entry.get()}/10')
            textresult.insert(END, f'\nTravail pratique 2 : {TP2Entry.get()}/10')
            textresult.insert(END, f'\nTravail pratique 3 : {TP3Entry.get()}/10')
            textresult.insert(END, f'\nTravail pratique 4 : {TP4Entry.get()}/10')
            textresult.insert(END, f'\nTravail pratique 5 : {TP5Entry.get()}/10')
            textresult.insert(END, f'\n                                         - - - - - - - - -')
            textresult.insert(END, f'\n                                         Moyenne : {totaltp.get()}/10')
            textresult.insert(END, f'\n                                         - - - - - - - - -')

            textresult.insert(END, f'\n\n==========================================================')
            textresult.insert(END, f'\nTravail journalier 1 : {TJ1Entry.get()}/10')
            textresult.insert(END, f'\nTravail journalier 2 : {TJ2Entry.get()}/10')
            textresult.insert(END, f'\nTravail journalier 3 : {TJ3Entry.get()}/10')
            textresult.insert(END, f'\nTravail journalier 4 : {TJ4Entry.get()}/10')
            textresult.insert(END, f'\nTravail journalier 5 : {TJ5Entry.get()}/10')
            textresult.insert(END, f'\n                                         - - - - - - - - -')
            textresult.insert(END, f'\n                                         Moyenne : {totaltj.get()}/10')
            textresult.insert(END, f'\n                                         - - - - - - - - -')

            textresult.insert(END, f'\n\n...........................................................\n')
            textresult.insert(END, f'\n                 INTERRO GENERAUX ET EXAMENS            ')
            textresult.insert(END, f'\n                 ---------------------------    ')
            textresult.insert(END, f'\nInterrogation générale période 1 : {IG1Entry.get()}/10')
            textresult.insert(END, f'\nInterrogation générale période 2 : {IG2Entry.get()}/10')
            textresult.insert(END, f'\nInterrogation générale période 3 : {IG3Entry.get()}/10')
            textresult.insert(END, f'\nExamen sémestre 1                : {Exa1Entry.get()}/10')
            textresult.insert(END, f'\nExamen sémestre 2                : {Exa2Entry.get()}/10')

            textresult.insert(END, f'\n\n==========================================================')
            textresult.insert(END, f'\n\nMoyenne annuelle : {frame.get()}/20')
            textresult.insert(END, f'\nPourcentage      : {pourcentage.get()}')
            textresult.insert(END, f'\nGrade            : {Grade.get()}')
            if Grade == "L":
                textresult.insert(END, f"\n\nDécision : Echec grave")
            if Grade.get() == "F":
                textresult.insert(END, f"\n\nDécision : Insuffisant")
            if Grade.get() == "G":
                textresult.insert(END, f"\n\nDécision : Assez bien")
            if Grade.get() == "B":
                textresult.insert(END, f"\n\nDécision : Bien")
            if Grade.get() == "D":
                textresult.insert(END, f"\n\nDécision : Très bien")
            if Grade.get() == "E":
                textresult.insert(END, f"\n\nDécision : Excellent")
            textresult.insert(END, f'\n                                      ')
            textresult.insert(END, f'\n~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~')
            textresult.insert(END, f'\n                                            Copyright@StimLog2023')

    def reinitialiser():
        entryname.delete(0, END)
        entrynumber.delete(0, END)
        entryCOURS.delete(0, END)
        entrymatricule.delete(0, END)

        interro1Entry.delete(0, END)
        interro2Entry.delete(0, END)
        interro3Entry.delete(0, END)
        interro4Entry.delete(0, END)
        interro5Entry.delete(0, END)
        TP1Entry.delete(0, END)
        TP2Entry.delete(0, END)
        TP3Entry.delete(0, END)
        TP4Entry.delete(0, END)
        TP5Entry.delete(0, END)
        TJ1Entry.delete(0, END)
        TJ2Entry.delete(0, END)
        TJ3Entry.delete(0, END)
        TJ4Entry.delete(0, END)
        TJ5Entry.delete(0, END)
        IG1Entry.delete(0, END)
        IG2Entry.delete(0, END)
        IG3Entry.delete(0, END)
        Exa1Entry.delete(0, END)
        Exa2Entry.delete(0, END)

        interro1Entry.insert(0, 0)
        interro2Entry.insert(0, 0)
        interro3Entry.insert(0, 0)
        interro4Entry.insert(0, 0)
        interro5Entry.insert(0, 0)
        TP1Entry.insert(0, 0)
        TP2Entry.insert(0, 0)
        TP3Entry.insert(0, 0)
        TP4Entry.insert(0, 0)
        TP5Entry.insert(0, 0)
        TJ1Entry.insert(0, 0)
        TJ2Entry.insert(0, 0)
        TJ3Entry.insert(0, 0)
        TJ4Entry.insert(0, 0)
        TJ5Entry.insert(0, 0)
        IG1Entry.insert(0, 0)
        IG2Entry.insert(0, 0)
        IG3Entry.insert(0, 0)
        Exa1Entry.insert(0, 0)
        Exa2Entry.insert(0, 0)

        totalInterrogation.delete(0, END)
        totaltp.delete(0, END)
        totaltj.delete(0, END)
        frame.delete(0, END)
        pourcentage.delete(0, END)
        Grade.delete(0, END)

        textresult.delete(1.0, END)

    def quitter():
        question = messagebox.askyesno('Exit', 'Voulez-vous fermer le programme ?')
        if question:
            quit()

    # FENETRE PRINCIPALE
    root = Tk()
    root.config(background='#0f1520')
    root.geometry('1290x650')
    root.minsize(1290, 650)
    root.maxsize(1290, 650)
    root.title('StimLink Logiciel')
    root.iconbitmap('Design StimLink 2-1.ico')

    labelprincipale = Label(root, text="StimLink Logiciel", font=('Arial', 30, 'bold'), bg='#0f1520', fg='white',
                            bd=12, relief=RIDGE)
    labelprincipale.pack(fill=X)

    detail_frame = LabelFrame(root, text='Détails élève', font=('Arial', 15, 'bold'),
                              fg='gold', bg='#0f1520', relief=GROOVE, bd=5)

    detail_frame.pack(fill=X)

    namelabel = Label(detail_frame, text='Nom α post-nom', font=('Arial', 13, 'bold'), bg='#0f1520', fg='white')
    namelabel.grid(row=0, column=0, padx=16, pady=2)
    entryname = Entry(detail_frame, font=('Arial', 13, 'bold'), bd=4, width=17)
    entryname.grid(row=0, column=1, padx=8)

    numberlabel = Label(detail_frame, text='Classe', font=('Arial', 13, 'bold'), bg='#0f1520', fg='white')
    numberlabel.grid(row=0, column=2, padx=10, pady=2)
    ListClasse = ['8ème A', '8ème B', '8ème C', '1ère ETRI A', '1ère ETRI B', '1ère ETRI C',
                  '2ème ETRI A', '2ème ETRI B', '2ème ETRI C', '3ème ETRI A', '3ème ETRI B',
                  '3ème ETRI C', '4ème ETRI A', '4ème ETRI B', '4ème ETRI C']
    entrynumber = ttk.Combobox(detail_frame, values=ListClasse, font=('Arial', 11, 'bold'), width=11)
    entrynumber.grid(row=0, column=3, padx=8)

    COURSlabel = Label(detail_frame, text='Cours', font=('Arial', 13, 'bold'), bg='#0f1520', fg='white')
    COURSlabel.grid(row=0, column=4, padx=15, pady=2)
    listCours = ['Applications', 'Dessin schéma 1', 'Dessin schéma 2', 'Dissertation', 'Electricité 1',
                 'Electricité 2', 'Electricité 3', 'Informatique 1', 'Information 5', 'Informatique 6']
    entryCOURS = ttk.Combobox(detail_frame, values=listCours, font=('Arial', 12, 'bold'), width=15)
    entryCOURS.grid(row=0, column=5, padx=8)

    matriculelabel = Label(detail_frame, text='Matricule', font=('Arial', 13, 'bold'), bg='#0f1520', fg='blue')
    matriculelabel.grid(row=0, column=6, padx=6, pady=2)
    entrymatricule = Entry(detail_frame, font=('Arial', 13, 'bold'), bd=4, width=15)
    entrymatricule.grid(row=0, column=7, padx=8)

    Recherche = Button(detail_frame, text='Recherche', font=('Arial', 11, 'bold'), bd=3, width=8, command=recherche)
    Recherche.grid(row=0, column=8, padx=16, pady=5)

    BD = Button(detail_frame, text='Database', font=('Arial', 11, 'bold'), bd=3, bg='seagreen', fg='white', width=8,
                command=Base_de_donnees)
    BD.grid(row=0, column=9, padx=13, pady=5)

    eleveframe = Frame(root)
    eleveframe.pack()

    # Insertions
    interroframe = LabelFrame(eleveframe, text='Interrogations', font=('Arial', 15, 'bold'),
                              fg='gold', bg='#0f1520', relief=GROOVE, bd=5)
    interroframe.grid(row=0, column=0)

    newframe = Label(interroframe, text='Interro 1', font=('Arial', 13, 'bold'),
                     fg='white', bg='#0f1520')
    newframe.grid(row=0, column=0)

    interro1Entry = Entry(interroframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    interro1Entry.grid(row=0, column=1, pady=10, padx=10, sticky='w')
    interro1Entry.insert(0, 0)

    newframe1 = Label(interroframe, text='Interro 2', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe1.grid(row=1, column=0)

    interro2Entry = Entry(interroframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    interro2Entry.grid(row=1, column=1, pady=10, padx=10, sticky='w')
    interro2Entry.insert(0, 0)

    newframe2 = Label(interroframe, text='Interro 3', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe2.grid(row=2, column=0)

    interro3Entry = Entry(interroframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    interro3Entry.grid(row=2, column=1, pady=10, padx=10, sticky='w')
    interro3Entry.insert(0, 0)

    newframe3 = Label(interroframe, text='Interro 4', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe3.grid(row=3, column=0)

    interro4Entry = Entry(interroframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    interro4Entry.grid(row=3, column=1, pady=10, padx=10, sticky='w')
    interro4Entry.insert(0, 0)

    newframe4 = Label(interroframe, text='Interro 5', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe4.grid(row=4, column=0)

    interro5Entry = Entry(interroframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    interro5Entry.grid(row=4, column=1, pady=10, padx=10, sticky='w')
    interro5Entry.insert(0, 0)

    TPframe = LabelFrame(eleveframe, text='Travaux pratiques', font=('Arial', 15, 'bold'),
                         fg='gold', bg='#0f1520', relief=GROOVE, bd=5)
    TPframe.grid(row=0, column=1)

    newframe = Label(TPframe, text='TP 1', font=('Arial', 13, 'bold'),
                     fg='white', bg='#0f1520')
    newframe.grid(row=0, column=0)

    TP1Entry = Entry(TPframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    TP1Entry.grid(row=0, column=1, pady=10, padx=10, sticky='w')
    TP1Entry.insert(0, 0)
    newframe2 = Label(TPframe, text='TP 2', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe2.grid(row=1, column=0)

    TP2Entry = Entry(TPframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    TP2Entry.grid(row=1, column=1, pady=10, padx=10, sticky='w')
    TP2Entry.insert(0, 0)

    newframe3 = Label(TPframe, text='TP 3', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe3.grid(row=2, column=0)

    TP3Entry = Entry(TPframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    TP3Entry.grid(row=2, column=1, pady=10, padx=10, sticky='w')
    TP3Entry.insert(0, 0)

    newframe4 = Label(TPframe, text='TP 4', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe4.grid(row=3, column=0)

    TP4Entry = Entry(TPframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    TP4Entry.grid(row=3, column=1, pady=10, padx=10, sticky='w')
    TP4Entry.insert(0, 0)

    newframe5 = Label(TPframe, text='TP 5', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe5.grid(row=4, column=0)

    TP5Entry = Entry(TPframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    TP5Entry.grid(row=4, column=1, pady=10, padx=10, sticky='w')
    TP5Entry.insert(0, 0)

    TJframe = LabelFrame(eleveframe, text='Travaux journaliers', font=('Arial', 15, 'bold'),
                         fg='gold', bg='#0f1520', relief=GROOVE, bd=5)
    TJframe.grid(row=0, column=2)

    newframe = Label(TJframe, text='TJ 1', font=('Arial', 13, 'bold'),
                     fg='white', bg='#0f1520')
    newframe.grid(row=0, column=0)

    TJ1Entry = Entry(TJframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    TJ1Entry.grid(row=0, column=1, pady=10, padx=10, sticky='w')
    TJ1Entry.insert(0, 0)

    newframe1 = Label(TJframe, text='TJ 2', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe1.grid(row=1, column=0)

    TJ2Entry = Entry(TJframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    TJ2Entry.grid(row=1, column=1, pady=10, padx=10, sticky='w')
    TJ2Entry.insert(0, 0)

    newframe2 = Label(TJframe, text='TJ 3', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe2.grid(row=2, column=0)

    TJ3Entry = Entry(TJframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    TJ3Entry.grid(row=2, column=1, pady=10, padx=10, sticky='w')
    TJ3Entry.insert(0, 0)

    newframe3 = Label(TJframe, text='TJ 4', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe3.grid(row=3, column=0)

    TJ4Entry = Entry(TJframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    TJ4Entry.grid(row=3, column=1, pady=10, padx=10, sticky='w')
    TJ4Entry.insert(0, 0)

    newframe4 = Label(TJframe, text='TJ 5', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe4.grid(row=4, column=0)

    TJ5Entry = Entry(TJframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    TJ5Entry.grid(row=4, column=1, pady=10, padx=10, sticky='w')
    TJ5Entry.insert(0, 0)

    TJframe = LabelFrame(eleveframe, text='Examens et IG', font=('Arial', 15, 'bold'),
                         fg='gold', bg='#0f1520', relief=GROOVE, bd=5)
    TJframe.grid(row=0, column=3, sticky='w')

    newframe = Label(TJframe, text='IG 1', font=('Arial', 13, 'bold'),
                     fg='white', bg='#0f1520')
    newframe.grid(row=0, column=0, sticky='w')

    IG1Entry = Entry(TJframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    IG1Entry.grid(row=0, column=1, pady=10, padx=10, sticky='w')
    IG1Entry.insert(0, 0)

    newframe1 = Label(TJframe, text='IG 2', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe1.grid(row=1, column=0, sticky='w')

    IG2Entry = Entry(TJframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    IG2Entry.grid(row=1, column=1, pady=10, padx=10)
    IG2Entry.insert(0, 0)

    newframe2 = Label(TJframe, text='IG 3', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe2.grid(row=2, column=0, sticky='w')

    IG3Entry = Entry(TJframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    IG3Entry.grid(row=2, column=1, pady=10, padx=10)
    IG3Entry.insert(0, 0)

    newframe2 = Label(TJframe, text='Exam 1', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe2.grid(row=3, column=0, sticky='w')

    Exa1Entry = Entry(TJframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    Exa1Entry.grid(row=3, column=1, pady=10, padx=10)
    Exa1Entry.insert(0, 0)

    newframe2 = Label(TJframe, text='Exam 2', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    newframe2.grid(row=4, column=0, sticky='w')

    Exa2Entry = Entry(TJframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    Exa2Entry.grid(row=4, column=1, pady=10, padx=10)
    Exa2Entry.insert(0, 0)

    billframe = Frame(eleveframe, bd=7, relief=GROOVE)
    billframe.grid(row=0, column=4, padx=10)

    billLabel = Label(billframe, text='BULLETIN DE RESULTAT', font=('Arial', 13, 'bold'), bd=3, relief=GROOVE)
    billLabel.pack(fill=X)

    scrollbar = Scrollbar(billframe, orient=VERTICAL)
    scrollbar.pack(side=RIGHT, fill=Y)

    textresult = Text(billframe, height=14, width=65, yscrollcommand=scrollbar.set)
    textresult.pack()
    scrollbar.config(command=textresult.yview())

    btnframe = LabelFrame(root, text='Moyennes générales', font=('Arial', 15, 'bold'),
                          fg='gold', bg='#0f1520', relief=GROOVE, bd=5)
    btnframe.pack()

    Totale_interro = Label(btnframe, text='Totale interrogations', font=('Arial', 13, 'bold'),
                           fg='white', bg='#0f1520')
    Totale_interro.grid(row=0, column=0, sticky='w')

    totalInterrogation = Entry(btnframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    totalInterrogation.grid(row=0, column=1, pady=8, padx=9)

    Totale_tp = Label(btnframe, text='Totale travaux pratiques', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    Totale_tp.grid(row=1, column=0, sticky='w')

    totaltp = Entry(btnframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    totaltp.grid(row=1, column=1, pady=8, padx=9)

    Totale_tj = Label(btnframe, text='Totale travaux journaliers', font=('Arial', 13, 'bold'),
                      fg='white', bg='#0f1520')
    Totale_tj.grid(row=2, column=0, sticky='w')

    totaltj = Entry(btnframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    totaltj.grid(row=2, column=1, pady=8, padx=9)
    # -----------------------------------------------------
    Moyenne = Label(btnframe, text='Moyenne générale', font=('Arial', 13, 'bold'),
                    fg='white', bg='#0f1520')
    Moyenne.grid(row=0, column=3, sticky='w')

    frame = Entry(btnframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    frame.grid(row=0, column=4, pady=8, padx=9)

    pourcent = Label(btnframe, text='Pourcentage', font=('Arial', 13, 'bold'),
                     fg='white', bg='#0f1520')
    pourcent.grid(row=1, column=3, sticky='w')

    pourcentage = Entry(btnframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    pourcentage.grid(row=1, column=4, pady=8, padx=9)

    grade = Label(btnframe, text='Grade', font=('Arial', 13, 'bold'),
                  fg='white', bg='#0f1520')
    grade.grid(row=2, column=3, sticky='w')

    Grade = Entry(btnframe, font=('Arial', 13, 'bold'), width=6, bd=3)
    Grade.grid(row=2, column=4, pady=8, padx=9)

    # BOUTONS
    boutonsFrame = Frame(btnframe, bd=4, relief=GROOVE)
    boutonsFrame.grid(row=0, column=5, rowspan=3)

    sommeBouton = Button(boutonsFrame, text='Calcule', font=('arial', 12, 'bold'), bg='#0f1520', fg='white', bd=2,
                         width=7, pady=7, command=somme)
    sommeBouton.grid(row=0, column=0, pady=29, padx=5)

    ApercuBouton = Button(boutonsFrame, text='Aperçu', font=('arial', 12, 'bold'), bg='#0f1520', fg='white', bd=2,
                          width=7, pady=7, command=apercu)
    ApercuBouton.grid(row=0, column=1, pady=15, padx=8)

    ImpressionBouton = Button(boutonsFrame, text='Sauvegarder', font=('arial', 12, 'bold'), bg='#0f1520', fg='white',
                              bd=2, width=10, pady=7, command=sauvegarder)
    ImpressionBouton.grid(row=0, column=2, pady=15, padx=8)

    SupprimerBouton = Button(boutonsFrame, text='E-mail', font=('arial', 12, 'bold'), bg='#0f1520', fg='white', bd=2,
                             width=6, pady=7, command=email)
    SupprimerBouton.grid(row=0, column=3, pady=15, padx=8)

    DataBouton = Button(boutonsFrame, text='Imprimer', font=('arial', 12, 'bold'), bg='#0f1520', fg='white', bd=2,
                        width=9, pady=7, command=imprimer)
    DataBouton.grid(row=0, column=4, pady=15, padx=8)

    RBouton = Button(boutonsFrame, text='Réinitialiser', font=('arial', 12, 'bold'), bg='#0f1520', fg='white',
                     bd=2, width=10, pady=7, command=reinitialiser)
    RBouton.grid(row=0, column=5, pady=15, padx=8)

    QuitterBouton = Button(boutonsFrame, text='Quitter', font=('arial', 12, 'bold'), bg='red', fg='white',
                           bd=3, width=7, pady=6, command=quitter)
    QuitterBouton.grid(row=0, column=6, pady=13, padx=7)

    root.mainloop()


def destroy():
    global fen
    fen.destroy()


accesibilite()
